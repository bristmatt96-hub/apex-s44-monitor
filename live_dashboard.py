#!/usr/bin/env python3
"""
Live Dashboard — auto-refreshes spreadsheet in Numbers with live API data.

Usage:
    python live_dashboard.py            # Refresh every 30 seconds
    python live_dashboard.py 60         # Refresh every 60 seconds
"""
import sys
import os
import time
import json
import subprocess
import urllib.request
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

API_BASE = "https://app.mb-trading.co.uk"
XLSX_PATH = "/Users/matt/Projects/credit-catalyst/APEX_Trading_Dashboard.xlsx"

# --- Colors ---
DARK_BG = "1A1A2E"
CARD_BG = "16213E"
HEADER_BG = "0F3460"
GREEN = "00C853"
RED = "FF1744"
BLUE = "448AFF"
MUTED = "8899AA"
WHITE = "FFFFFF"
YELLOW = "FFD600"

white_font = Font(name="Calibri", size=11, color=WHITE)
white_bold = Font(name="Calibri", bold=True, size=11, color=WHITE)
muted_font = Font(name="Calibri", size=10, color=MUTED)
card_fill = PatternFill(start_color=CARD_BG, end_color=CARD_BG, fill_type="solid")
dark_fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
thin_border = Border(
    left=Side(style="thin", color=MUTED),
    right=Side(style="thin", color=MUTED),
    top=Side(style="thin", color=MUTED),
    bottom=Side(style="thin", color=MUTED),
)


def fetch_json(endpoint):
    try:
        url = f"{API_BASE}{endpoint}"
        req = urllib.request.Request(url, headers={"Accept": "application/json"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            return json.loads(resp.read().decode())
    except Exception as e:
        print(f"  [WARN] Failed to fetch {endpoint}: {e}")
        return None


def pnl_color(value, big=False):
    size = 24 if big else 12
    color = GREEN if value >= 0 else RED
    return Font(name="Calibri", bold=True, size=size, color=color)


def refresh_file():
    """Write updated data to the xlsx file"""
    now = datetime.now().strftime("%d %b %Y %H:%M:%S")

    pnl = fetch_json("/api/pnl")
    positions = fetch_json("/api/positions")
    opportunities = fetch_json("/api/opportunities")

    wb = openpyxl.load_workbook(XLSX_PATH)

    # === DASHBOARD ===
    ws = wb["Dashboard"]
    ws['B3'] = f"Last Updated: {now}  (LIVE)"
    ws['B3'].font = muted_font
    ws['B3'].fill = dark_fill

    if pnl:
        ws['C6'] = pnl['daily_pnl']
        ws['C6'].font = pnl_color(pnl['daily_pnl'], big=True)
        ws['C6'].number_format = '[Green]+$#,##0.00;[Red]-$#,##0.00'
        ws['C6'].fill = dark_fill

        ws['C7'] = pnl['daily_pnl_pct'] / 100
        ws['C7'].font = pnl_color(pnl['daily_pnl_pct'])
        ws['C7'].number_format = '+0.00%;-0.00%'
        ws['C7'].fill = dark_fill

        ws['F6'] = pnl['ytd_pnl']
        ws['F6'].font = pnl_color(pnl['ytd_pnl'], big=True)
        ws['F6'].number_format = '[Green]+$#,##0.00;[Red]-$#,##0.00'
        ws['F6'].fill = dark_fill

        ws['F7'] = pnl['ytd_pnl_pct'] / 100
        ws['F7'].font = pnl_color(pnl['ytd_pnl_pct'])
        ws['F7'].number_format = '+0.00%;-0.00%'
        ws['F7'].fill = dark_fill

        ws['C9'] = pnl['total_positions']
        ws['C9'].font = white_bold
        ws['C9'].fill = card_fill

        ws['E9'] = pnl['winning_positions']
        ws['E9'].font = Font(name="Calibri", bold=True, size=11, color=GREEN)
        ws['E9'].fill = card_fill

        ws['G9'] = pnl['losing_positions']
        ws['G9'].font = Font(name="Calibri", bold=True, size=11, color=RED)
        ws['G9'].fill = card_fill

        ws['I9'] = pnl['realized_today']
        ws['I9'].font = pnl_color(pnl['realized_today'])
        ws['I9'].number_format = '[Green]+$#,##0.00;[Red]-$#,##0.00'
        ws['I9'].fill = card_fill

    if positions is not None:
        for row in range(14, 24):
            for col in range(2, 12):
                cell = ws.cell(row=row, column=col)
                cell.value = None
                cell.fill = card_fill
                cell.font = white_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

        for i, pos in enumerate(positions[:10]):
            row = 14 + i
            ws.cell(row=row, column=2).value = pos['symbol']
            ws.cell(row=row, column=2).font = white_bold
            ws.cell(row=row, column=3).value = pos['market_type'].upper()
            ws.cell(row=row, column=3).font = muted_font
            ws.cell(row=row, column=4).value = pos['quantity']
            ws.cell(row=row, column=5).value = pos['entry_price']
            ws.cell(row=row, column=5).number_format = '$#,##0.00'
            ws.cell(row=row, column=6).value = pos['current_price']
            ws.cell(row=row, column=6).number_format = '$#,##0.00'
            ws.cell(row=row, column=7).value = pos['unrealized_pnl']
            ws.cell(row=row, column=7).font = pnl_color(pos['unrealized_pnl'])
            ws.cell(row=row, column=7).number_format = '[Green]+$#,##0.00;[Red]-$#,##0.00'
            ws.cell(row=row, column=8).value = pos['unrealized_pnl_pct'] / 100
            ws.cell(row=row, column=8).font = pnl_color(pos['unrealized_pnl_pct'])
            ws.cell(row=row, column=8).number_format = '[Green]+0.00%;[Red]-0.00%'
            ws.cell(row=row, column=9).value = pos.get('composite_score') or 0
            ws.cell(row=row, column=9).font = white_bold
            ws.cell(row=row, column=10).value = pos.get('strategy') or '-'
            ws.cell(row=row, column=10).font = muted_font
            sl = pos.get('stop_loss')
            ws.cell(row=row, column=11).value = sl if sl else '-'
            ws.cell(row=row, column=11).font = Font(name="Calibri", size=11, color=RED) if sl else muted_font
            if sl:
                ws.cell(row=row, column=11).number_format = '$#,##0.00'
            for col in range(2, 12):
                ws.cell(row=row, column=col).fill = card_fill
                ws.cell(row=row, column=col).border = thin_border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")

    # === OPPORTUNITIES ===
    ws2 = wb["Opportunities"]
    ws2['B3'] = f"Last Updated: {now}  (LIVE)"
    ws2['B3'].font = muted_font
    ws2['B3'].fill = dark_fill

    if opportunities is not None:
        for row in range(6, 16):
            for col in range(2, 13):
                cell = ws2.cell(row=row, column=col)
                cell.value = None
                cell.fill = card_fill
                cell.font = white_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

        for i, opp in enumerate(opportunities[:10]):
            row = 6 + i
            ws2.cell(row=row, column=2).value = opp['rank']
            ws2.cell(row=row, column=2).font = Font(name="Calibri", bold=True, size=11, color=BLUE)
            ws2.cell(row=row, column=3).value = opp['symbol']
            ws2.cell(row=row, column=3).font = white_bold
            ws2.cell(row=row, column=4).value = opp['market_type'].upper()
            ws2.cell(row=row, column=4).font = muted_font
            score = opp['composite_score']
            score_color = GREEN if score >= 8 else (YELLOW if score >= 6 else RED)
            ws2.cell(row=row, column=5).value = score
            ws2.cell(row=row, column=5).font = Font(name="Calibri", bold=True, size=11, color=score_color)
            ws2.cell(row=row, column=6).value = f"{opp['risk_reward']:.1f}:1"
            ws2.cell(row=row, column=6).font = Font(name="Calibri", bold=True, size=11, color=GREEN)
            ws2.cell(row=row, column=7).value = f"{opp['confidence'] * 100:.0f}%"
            ws2.cell(row=row, column=7).font = white_bold
            ws2.cell(row=row, column=8).value = opp['entry_price']
            ws2.cell(row=row, column=8).number_format = '$#,##0.00'
            ws2.cell(row=row, column=9).value = opp['target_price']
            ws2.cell(row=row, column=9).font = Font(name="Calibri", size=11, color=GREEN)
            ws2.cell(row=row, column=9).number_format = '$#,##0.00'
            ws2.cell(row=row, column=10).value = opp['stop_loss']
            ws2.cell(row=row, column=10).font = Font(name="Calibri", size=11, color=RED)
            ws2.cell(row=row, column=10).number_format = '$#,##0.00'
            ws2.cell(row=row, column=11).value = opp.get('strategy') or '-'
            ws2.cell(row=row, column=11).font = muted_font
            reasoning = opp.get('reasoning', [])
            ws2.cell(row=row, column=12).value = " | ".join(reasoning) if reasoning else '-'
            ws2.cell(row=row, column=12).font = muted_font
            ws2.cell(row=row, column=12).alignment = Alignment(horizontal="left", wrap_text=True)
            for col in range(2, 13):
                ws2.cell(row=row, column=col).fill = card_fill
                ws2.cell(row=row, column=col).border = thin_border

    wb.save(XLSX_PATH)
    return pnl, positions, opportunities


def reload_numbers():
    """Tell Numbers to close and reopen the file to show updated data"""
    script = f'''
    tell application "Numbers"
        set theDoc to missing value
        repeat with d in documents
            if name of d contains "APEX_Trading_Dashboard" then
                set theDoc to d
                exit repeat
            end if
        end repeat
        if theDoc is not missing value then
            close theDoc saving no
        end if
        open POSIX file "{XLSX_PATH}"
    end tell
    '''
    subprocess.run(["osascript", "-e", script], capture_output=True)


def main():
    interval = int(sys.argv[1]) if len(sys.argv) > 1 else 30

    print(f"APEX Live Dashboard")
    print(f"Refreshing every {interval}s. Press Ctrl+C to stop.\n")

    # Open in Numbers initially
    subprocess.run(["open", "-a", "Numbers", XLSX_PATH], capture_output=True)
    time.sleep(3)

    try:
        while True:
            now = datetime.now().strftime("%H:%M:%S")
            print(f"[{now}] Refreshing...", end="", flush=True)
            try:
                pnl, positions, opportunities = refresh_file()
                reload_numbers()
                parts = []
                if positions is not None:
                    parts.append(f"Pos: {len(positions)}")
                if opportunities is not None:
                    parts.append(f"Opp: {len(opportunities)}")
                if pnl:
                    parts.append(f"Day: ${pnl['daily_pnl']:+.2f}")
                    parts.append(f"YTD: ${pnl['ytd_pnl']:+.2f}")
                print(f"  {' | '.join(parts)}")
            except Exception as e:
                print(f"  [ERROR] {e}")
            time.sleep(interval)
    except KeyboardInterrupt:
        print("\nStopped.")


if __name__ == "__main__":
    main()
