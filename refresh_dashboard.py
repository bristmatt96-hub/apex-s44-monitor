#!/usr/bin/env python3
"""
Refresh Excel dashboard with live data from the trading API.

Usage:
    python refresh_dashboard.py              # One-time refresh
    python refresh_dashboard.py --watch      # Auto-refresh every 30 seconds
    python refresh_dashboard.py --watch 60   # Auto-refresh every 60 seconds
"""
import sys
import time
import json
import urllib.request
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

API_BASE = "https://app.mb-trading.co.uk"
XLSX_PATH = "/Users/matt/Projects/credit-catalyst/APEX_Trading_Dashboard.xlsx"

# --- Colors ---
DARK_BG = "1A1A2E"
CARD_BG = "16213E"
HEADER_BG = "0F3460"
GREEN = "00C853"
RED = "FF1744"
BLUE = "448AFF"
MUTED = "8899AA"
WHITE = "FFFFFF"
YELLOW = "FFD600"

header_font = Font(name="Calibri", bold=True, size=11, color=WHITE)
header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
card_fill = PatternFill(start_color=CARD_BG, end_color=CARD_BG, fill_type="solid")
dark_fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
white_font = Font(name="Calibri", size=11, color=WHITE)
white_bold = Font(name="Calibri", bold=True, size=11, color=WHITE)
muted_font = Font(name="Calibri", size=10, color=MUTED)
title_font = Font(name="Calibri", bold=True, size=16, color=WHITE)
subtitle_font = Font(name="Calibri", bold=True, size=13, color=BLUE)
thin_border = Border(
    left=Side(style="thin", color=MUTED),
    right=Side(style="thin", color=MUTED),
    top=Side(style="thin", color=MUTED),
    bottom=Side(style="thin", color=MUTED),
)


def fetch_json(endpoint):
    """Fetch JSON from API endpoint"""
    try:
        url = f"{API_BASE}{endpoint}"
        req = urllib.request.Request(url, headers={"Accept": "application/json"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            return json.loads(resp.read().decode())
    except Exception as e:
        print(f"  [WARN] Failed to fetch {endpoint}: {e}")
        return None


def pnl_color(value):
    """Return green or red font based on value"""
    if value >= 0:
        return Font(name="Calibri", bold=True, size=12, color=GREEN)
    return Font(name="Calibri", bold=True, size=12, color=RED)


def pnl_color_big(value):
    if value >= 0:
        return Font(name="Calibri", bold=True, size=24, color=GREEN)
    return Font(name="Calibri", bold=True, size=24, color=RED)


def refresh():
    """Fetch live data and update the Excel spreadsheet"""
    now = datetime.now().strftime("%d %b %Y %H:%M:%S")
    print(f"[{now}] Refreshing dashboard...")

    # Fetch all data
    pnl = fetch_json("/api/pnl")
    positions = fetch_json("/api/positions")
    opportunities = fetch_json("/api/opportunities")

    # Load workbook
    wb = openpyxl.load_workbook(XLSX_PATH)

    # ===========================
    # SHEET 1: DASHBOARD
    # ===========================
    ws = wb["Dashboard"]

    # Update timestamp
    ws['B3'] = f"Last Updated: {now}  (Live from API)"
    ws['B3'].font = muted_font
    ws['B3'].fill = dark_fill

    if pnl:
        # Daily P&L
        ws['C6'] = pnl['daily_pnl']
        ws['C6'].font = pnl_color_big(pnl['daily_pnl'])
        ws['C6'].number_format = '[Green]+$#,##0.00;[Red]-$#,##0.00'
        ws['C6'].fill = dark_fill

        ws['C7'] = pnl['daily_pnl_pct'] / 100
        ws['C7'].font = pnl_color(pnl['daily_pnl_pct'])
        ws['C7'].number_format = '+0.00%;-0.00%'
        ws['C7'].fill = dark_fill

        # YTD P&L
        ws['F6'] = pnl['ytd_pnl']
        ws['F6'].font = pnl_color_big(pnl['ytd_pnl'])
        ws['F6'].number_format = '[Green]+$#,##0.00;[Red]-$#,##0.00'
        ws['F6'].fill = dark_fill

        ws['F7'] = pnl['ytd_pnl_pct'] / 100
        ws['F7'].font = pnl_color(pnl['ytd_pnl_pct'])
        ws['F7'].number_format = '+0.00%;-0.00%'
        ws['F7'].fill = dark_fill

        # Quick stats
        ws['C9'] = pnl['total_positions']
        ws['C9'].font = white_bold
        ws['C9'].fill = card_fill

        ws['E9'] = pnl['winning_positions']
        ws['E9'].font = Font(name="Calibri", bold=True, size=11, color=GREEN)
        ws['E9'].fill = card_fill

        ws['G9'] = pnl['losing_positions']
        ws['G9'].font = Font(name="Calibri", bold=True, size=11, color=RED)
        ws['G9'].fill = card_fill

        ws['I9'] = pnl['realized_today']
        ws['I9'].font = pnl_color(pnl['realized_today'])
        ws['I9'].number_format = '[Green]+$#,##0.00;[Red]-$#,##0.00'
        ws['I9'].fill = card_fill

        print(f"  P&L: Daily ${pnl['daily_pnl']:+.2f} | YTD ${pnl['ytd_pnl']:+.2f}")

    if positions is not None:
        # Clear old position rows
        for row in range(14, 24):
            for col in range(2, 12):
                cell = ws.cell(row=row, column=col)
                cell.value = None
                cell.fill = card_fill
                cell.font = white_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

        # Fill positions
        for i, pos in enumerate(positions[:10]):
            row = 14 + i
            ws.cell(row=row, column=2).value = pos['symbol']
            ws.cell(row=row, column=2).font = white_bold
            ws.cell(row=row, column=2).fill = card_fill

            ws.cell(row=row, column=3).value = pos['market_type'].upper()
            ws.cell(row=row, column=3).font = muted_font
            ws.cell(row=row, column=3).fill = card_fill

            ws.cell(row=row, column=4).value = pos['quantity']
            ws.cell(row=row, column=4).font = white_font
            ws.cell(row=row, column=4).fill = card_fill

            ws.cell(row=row, column=5).value = pos['entry_price']
            ws.cell(row=row, column=5).font = white_font
            ws.cell(row=row, column=5).number_format = '$#,##0.00'
            ws.cell(row=row, column=5).fill = card_fill

            ws.cell(row=row, column=6).value = pos['current_price']
            ws.cell(row=row, column=6).font = white_font
            ws.cell(row=row, column=6).number_format = '$#,##0.00'
            ws.cell(row=row, column=6).fill = card_fill

            ws.cell(row=row, column=7).value = pos['unrealized_pnl']
            ws.cell(row=row, column=7).font = pnl_color(pos['unrealized_pnl'])
            ws.cell(row=row, column=7).number_format = '[Green]+$#,##0.00;[Red]-$#,##0.00'
            ws.cell(row=row, column=7).fill = card_fill

            ws.cell(row=row, column=8).value = pos['unrealized_pnl_pct'] / 100
            ws.cell(row=row, column=8).font = pnl_color(pos['unrealized_pnl_pct'])
            ws.cell(row=row, column=8).number_format = '[Green]+0.00%;[Red]-0.00%'
            ws.cell(row=row, column=8).fill = card_fill

            score = pos.get('composite_score', 0) or 0
            ws.cell(row=row, column=9).value = score
            ws.cell(row=row, column=9).font = white_bold
            ws.cell(row=row, column=9).fill = card_fill

            ws.cell(row=row, column=10).value = pos.get('strategy') or '-'
            ws.cell(row=row, column=10).font = muted_font
            ws.cell(row=row, column=10).fill = card_fill

            sl = pos.get('stop_loss')
            ws.cell(row=row, column=11).value = sl if sl else '-'
            ws.cell(row=row, column=11).font = Font(name="Calibri", size=11, color=RED) if sl else muted_font
            ws.cell(row=row, column=11).number_format = '$#,##0.00' if sl else '@'
            ws.cell(row=row, column=11).fill = card_fill

        print(f"  Positions: {len(positions)} open")

    # ===========================
    # SHEET 2: OPPORTUNITIES
    # ===========================
    ws2 = wb["Opportunities"]

    ws2['B3'] = f"Last Updated: {now}  (Live from API)"
    ws2['B3'].font = muted_font
    ws2['B3'].fill = dark_fill

    if opportunities is not None:
        # Clear old rows
        for row in range(6, 16):
            for col in range(2, 13):
                cell = ws2.cell(row=row, column=col)
                cell.value = None
                cell.fill = card_fill
                cell.font = white_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

        for i, opp in enumerate(opportunities[:10]):
            row = 6 + i
            ws2.cell(row=row, column=2).value = opp['rank']
            ws2.cell(row=row, column=2).font = Font(name="Calibri", bold=True, size=11, color=BLUE)
            ws2.cell(row=row, column=2).fill = card_fill

            ws2.cell(row=row, column=3).value = opp['symbol']
            ws2.cell(row=row, column=3).font = white_bold
            ws2.cell(row=row, column=3).fill = card_fill

            ws2.cell(row=row, column=4).value = opp['market_type'].upper()
            ws2.cell(row=row, column=4).font = muted_font
            ws2.cell(row=row, column=4).fill = card_fill

            score = opp['composite_score']
            score_color = GREEN if score >= 8 else (YELLOW if score >= 6 else RED)
            ws2.cell(row=row, column=5).value = score
            ws2.cell(row=row, column=5).font = Font(name="Calibri", bold=True, size=11, color=score_color)
            ws2.cell(row=row, column=5).fill = card_fill

            ws2.cell(row=row, column=6).value = f"{opp['risk_reward']:.1f}:1"
            ws2.cell(row=row, column=6).font = Font(name="Calibri", bold=True, size=11, color=GREEN)
            ws2.cell(row=row, column=6).fill = card_fill

            ws2.cell(row=row, column=7).value = f"{opp['confidence'] * 100:.0f}%"
            ws2.cell(row=row, column=7).font = white_bold
            ws2.cell(row=row, column=7).fill = card_fill

            ws2.cell(row=row, column=8).value = opp['entry_price']
            ws2.cell(row=row, column=8).font = white_font
            ws2.cell(row=row, column=8).number_format = '$#,##0.00'
            ws2.cell(row=row, column=8).fill = card_fill

            ws2.cell(row=row, column=9).value = opp['target_price']
            ws2.cell(row=row, column=9).font = Font(name="Calibri", size=11, color=GREEN)
            ws2.cell(row=row, column=9).number_format = '$#,##0.00'
            ws2.cell(row=row, column=9).fill = card_fill

            ws2.cell(row=row, column=10).value = opp['stop_loss']
            ws2.cell(row=row, column=10).font = Font(name="Calibri", size=11, color=RED)
            ws2.cell(row=row, column=10).number_format = '$#,##0.00'
            ws2.cell(row=row, column=10).fill = card_fill

            ws2.cell(row=row, column=11).value = opp.get('strategy') or '-'
            ws2.cell(row=row, column=11).font = muted_font
            ws2.cell(row=row, column=11).fill = card_fill

            reasoning = opp.get('reasoning', [])
            ws2.cell(row=row, column=12).value = " | ".join(reasoning) if reasoning else '-'
            ws2.cell(row=row, column=12).font = muted_font
            ws2.cell(row=row, column=12).fill = card_fill
            ws2.cell(row=row, column=12).alignment = Alignment(horizontal="left", wrap_text=True)

        print(f"  Opportunities: {len(opportunities)} ranked")

    # Save
    wb.save(XLSX_PATH)
    print(f"  Saved to: {XLSX_PATH}")
    print()


def main():
    if "--watch" in sys.argv:
        idx = sys.argv.index("--watch")
        interval = int(sys.argv[idx + 1]) if idx + 1 < len(sys.argv) else 30
        print(f"Auto-refreshing every {interval} seconds. Press Ctrl+C to stop.\n")
        try:
            while True:
                refresh()
                time.sleep(interval)
        except KeyboardInterrupt:
            print("\nStopped.")
    else:
        refresh()


if __name__ == "__main__":
    main()
